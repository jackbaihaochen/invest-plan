# -*- coding: utf-8 -*-
"""1亿日元计划 —— Google 表格工作簿生成器。

通过 Drive 连接器创建表格会丢失公式（已验证：连 =1+1 都变空），
所以这里生成 .xlsx，由 Google 表格自己的导入器读入。

用法:  python3 tools/build-sheet.py build/1oku.xlsx
"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import LineChart, BarChart, Reference

INK, BAND, ACCENT = "0F172A", "1E293B", "2563EB"
GOOD, BAD, MUTED = "047857", "B91C1C", "64748B"
EDIT = "FEF9C3"

YEN, PCT, MON, DAY = '#,##0"円"', '0.0%', 'yyyy"年"m"月"', 'yyyy/m/d'
MONTHS, HOLDS = 72, 60

S_DASH, S_HOLD, S_MON, S_CFG = "总览", "持仓明细", "月次记录", "设置"
S_CSV = "CSV"
CSV_N = 200   # CSV 貼り付け範囲の行数

GOAL   = f"'{S_CFG}'!$B$2"
TARGET = f"'{S_CFG}'!$B$3"
RATE   = f"'{S_CFG}'!$B$4"
BASE   = f"'{S_CFG}'!$B$5"
PACE0  = f"'{S_CFG}'!$B$6"
MRATE  = f"({RATE}/12)"

M_YM  = f"'{S_MON}'!$A$3:$A${MONTHS+2}"
M_VAL = f"'{S_MON}'!$B$3:$B${MONTHS+2}"
M_IN  = f"'{S_MON}'!$C$3:$C${MONTHS+2}"
M_PRI = f"'{S_MON}'!$D$3:$D${MONTHS+2}"
M_STK = f"'{S_MON}'!$G$3:$G${MONTHS+2}"
H_CAT = f"'{S_HOLD}'!$B$3:$B${HOLDS+2}"
H_ACC = f"'{S_HOLD}'!$C$3:$C${HOLDS+2}"
H_VAL = f"'{S_HOLD}'!$J$3:$J${HOLDS+2}"
H_CST = f"'{S_HOLD}'!$K$3:$K${HOLDS+2}"
# CSV 側の列: A種別 B銘柄コード C銘柄 D口座 E保有数量 I現在値 O時価評価額[円] Q評価損益[円]
C_NAME = f"'{S_CSV}'!$C$1:$C${CSV_N}"
C_ACC  = f"'{S_CSV}'!$D$1:$D${CSV_N}"
C_QTY  = f"'{S_CSV}'!$E$1:$E${CSV_N}"
C_PX   = f"'{S_CSV}'!$I$1:$I${CSV_N}"
C_VAL  = f"'{S_CSV}'!$O$1:$O${CSV_N}"
C_PL   = f"'{S_CSV}'!$Q$1:$Q${CSV_N}"
C_LBL  = f"'{S_CSV}'!$A$1:$A${CSV_N}"
C_LVAL = f"'{S_CSV}'!$B$1:$B${CSV_N}"
FX    = f"'{S_CFG}'!$B$8"

CATS = ["指数基金", "美股个股", "日本个股", "黄金",
        "现金·MMF", "加密货币", "持株会", "银行存款"]
ACCTS = ["NISAつみたて投資枠", "NISA成長投資枠", "特定口座", "一般口座", "其他"]

# 楽天証券「保有商品詳細」2026/08/29。末尾2つは CSV 側の (銘柄, 口座) = 突き合わせキー。
# 数量・価格・取得原価は CSV を貼れば上書きされ、貼っていなければこの値が残る。
# (名称, 类别, 账户, 币种, 代码, 数量, 倍率, 手动价, 成本, CSV銘柄, CSV口座)
NISA_T, NISA_G, TOKUTEI, OTHER = "NISAつみたて投資枠", "NISA成長投資枠", "特定口座", "其他"
JPY, USD = "日元", "美元"
RAKU_SP = "楽天・プラス・Ｓ＆Ｐ５００インデックス・ファンド(楽天・プラス・Ｓ＆Ｐ５００)"
ORUKAN  = "eMAXIS Slim 全世界株式(オール・カントリー)(オルカン)"
HOLDINGS = [
    ("楽天・プラス・S&P500",        "指数基金", NISA_T,  JPY, "", 1513390, 10000, 20176, 2200000, RAKU_SP, NISA_T),
    ("楽天・プラス・S&P500",        "指数基金", NISA_G,  JPY, "", 1935641, 10000, 20176, 2609723, RAKU_SP, "NISA成長投資枠"),
    ("eMAXIS Slim オルカン",        "指数基金", NISA_G,  JPY, "",  207794, 10000, 38842,  619402, ORUKAN, "NISA成長投資枠"),
    ("eMAXIS Slim オルカン",        "指数基金", NISA_T,  JPY, "",  165294, 10000, 38842,  600000, ORUKAN, NISA_T),
    ("iFreeNEXT FANG+",             "指数基金", NISA_G,  JPY, "",   39902, 10000,100206,  350000, "iFreeNEXT FANG+インデックス", "NISA成長投資枠"),
    ("eMAXIS Slim 米国株式(S&P500)", "指数基金", NISA_G, JPY, "",   35254, 10000, 45344,  100000, "eMAXIS Slim 米国株式(S&P500)", "NISA成長投資枠"),
    ("楽天・プラス・NASDAQ-100",     "指数基金", NISA_G, JPY, "",   10155, 10000, 18978,   20000, "楽天・プラス・NASDAQ-100インデックス・ファンド(楽天・プラス・NASDAQ-100)", "NISA成長投資枠"),
    ("NVDA エヌビディア",           "美股个股", NISA_G,  USD, "NASDAQ:NVDA", 17, 1, 217.55, 338145, "エヌビディア", "NISA成長投資枠"),
    ("GOOG アルファベットC",        "美股个股", NISA_G,  USD, "NASDAQ:GOOG", 10, 1, 342.88, 265576, "アルファベット クラスC", "NISA成長投資枠"),
    ("CRCL サークル",               "美股个股", TOKUTEI, USD, "NYSE:CRCL",   18, 1,  87.14, 273264, "サークル・インターネット・グループ", "特定"),
    ("MSFT マイクロソフト",         "美股个股", NISA_G,  USD, "NASDAQ:MSFT",  2, 1, 513.53, 127332, "マイクロソフト", "NISA成長投資枠"),
    ("MSFT マイクロソフト",         "美股个股", TOKUTEI, USD, "NASDAQ:MSFT",  1, 1, 513.53,  60687, "マイクロソフト", "特定"),
    ("AMZN アマゾン",               "美股个股", NISA_G,  USD, "NASDAQ:AMZN",  2, 1, 266.43,  63981, "アマゾン・ドット・コム", "NISA成長投資枠"),
    ("AMZN アマゾン",               "美股个股", TOKUTEI, USD, "NASDAQ:AMZN",  1, 1, 266.43,  23722, "アマゾン・ドット・コム", "特定"),
    ("AMD",                         "美股个股", NISA_G,  USD, "NASDAQ:AMD",   1, 1, 465.58,  19102, "アドバンスト・マイクロ・デバイス(AMD)", "NISA成長投資枠"),
    ("TSM タイワンセミ",            "美股个股", NISA_G,  USD, "NYSE:TSM",     1, 1, 417.52,  65928, "タイワン・セミコンダクター・マニュファクチャリング", "NISA成長投資枠"),
    ("AAPL アップル",               "美股个股", NISA_G,  USD, "NASDAQ:AAPL",  1, 1, 319.70,  25715, "アップル", "NISA成長投資枠"),
    ("CLS セレスティカ",            "美股个股", NISA_G,  USD, "NYSE:CLS",     1, 1, 298.70,  60368, "セレスティカ", "NISA成長投資枠"),
    ("VRT バーティブ",              "美股个股", NISA_G,  USD, "NYSE:VRT",     1, 1, 257.08,  54219, "バーティブ・ホールディングス", "NISA成長投資枠"),
    ("6787 メイコー",               "日本个股", NISA_G,  JPY, "TYO:6787",     3, 1,  18760, 108763, "メイコー", "NISA成長投資枠"),
    ("4385 メルカリ",               "日本个股", NISA_G,  JPY, "TYO:4385",     7, 1,   4617,  13825, "メルカリ", "NISA成長投資枠"),
    ("3492 ＭＩＲＡＲＴＨ (REIT)",   "日本个股", NISA_G,  JPY, "TYO:3492",    1, 1,  79000,  92300, "ＭＩＲＡＲＴＨ不動産投", "NISA成長投資枠"),
    ("金 現物 (g)",                 "黄金",     OTHER,   JPY, "", 72.24958, 1,  23452, 1245041, "金", "-"),
    ("425A ＧＸゴールド",           "黄金",     NISA_G,  JPY, "TYO:425A",   380, 1,  390.9,  148610, "ＧＸゴールド", "NISA成長投資枠"),
    ("楽天・米ドルMMF",             "现金·MMF", TOKUTEI, JPY, "",        3105, 100, 160.06,   4934, "ノーザン・トラスト・米ドル・リクイディティ・ファンド(楽天・米ドルMMF)", "特定"),
    ("楽天証券 預り金",             "现金·MMF", OTHER,   JPY, "",           1, 1,   3478,    3478, "@CASH", ""),
    ("bitFlyer BTC",                "加密货币", OTHER,   JPY, "CURRENCY:BTCJPY", 0.061912, 1, None, 893703, "", ""),
    ("野村 持株会",                 "持株会",   OTHER,   JPY, "",           1, 1,   None,    None, "", ""),
    ("ゆうちょ銀行",                "银行存款", OTHER,   JPY, "",           1, 1,   None,    None, "", ""),
]

# 取引履歴から集計した楽天証券への純入金（円）
PAST = [("2025-09", 250470), ("2025-10", -50000), ("2025-11", -250000),
        ("2025-12", -100000), ("2026-01", 250000), ("2026-02", -40000),
        ("2026-03", 150000), ("2026-04", 200000), ("2026-05", 155587),
        ("2026-06", 310000), ("2026-07", 110000), ("2026-08", 100000)]


def bar(expr, width=20):
    n = f'ROUND(MIN(1,MAX(0,{expr}))*{width},0)'
    return f'=IFERROR(REPT("█",{n})&REPT("░",{width}-{n}),"")'


def mlabel(m):
    return f'=IFERROR(INT({m}/12)&"年"&ROUND(MOD({m},12),0)&"个月","—")'


def put(ws, ref, v, *, font=None, fill=None, fmt=None, align=None):
    c = ws[ref]; c.value = v
    if font: c.font = font
    if fill: c.fill = PatternFill("solid", fgColor=fill)
    if fmt:  c.number_format = fmt
    if align: c.alignment = align
    return c


def band(ws, row, text, first="B", last="G"):
    ws.merge_cells(f"{first}{row}:{last}{row}")
    put(ws, f"{first}{row}", text, font=Font(bold=True, size=11, color="FFFFFF"),
        fill=BAND, align=Alignment(vertical="center", indent=1))
    ws.row_dimensions[row].height = 24


def label(ws, row, text, col="B"):
    put(ws, f"{col}{row}", text, font=Font(size=10, color=MUTED),
        align=Alignment(vertical="center", indent=1))


def header(ws, row, names, widths):
    for col, w in zip(range(1, len(widths) + 1), widths):
        ws.column_dimensions[get_column_letter(col)].width = w
    for j, h in enumerate(names, start=1):
        put(ws, f"{get_column_letter(j)}{row}", h,
            font=Font(bold=True, size=10, color="FFFFFF"), fill=BAND,
            align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[row].height = 22


wb = Workbook()

# ------------------------------------------------------------------ 设置 --
cfg = wb.active; cfg.title = S_CFG
cfg.sheet_view.showGridLines = False
for c, w in zip("ABC", (26, 18, 60)):
    cfg.column_dimensions[c].width = w
put(cfg, "A1", "设置", font=Font(bold=True, size=16, color=INK))
for i, (name, val, fmt, note) in enumerate([
    ("目标金额",       100_000_000, YEN, "1亿日元。"),
    ("每月目标投入",       400_000, YEN, "按这个金额投，约10年到。"),
    ("预期年化收益",          0.07, PCT, "S&P500 长期平均。不是保证。"),
    ("起始累计本金",     9_301_761, YEN, "楽天証券取得原価 9,494,115 + BTC 893,703 − 已列在月次记录里的近12个月。还要加上野村持株会 / ゆうちょ 的本金。"),
    ("当前实际速度",        90_505, YEN, "楽天証券 近12个月净入金的月均（含取出的月份）。月次记录满3个月后自动改用实际值。"),
], start=2):
    put(cfg, f"A{i}", name, font=Font(size=11, color=INK), align=Alignment(indent=1))
    put(cfg, f"B{i}", val, font=Font(bold=True, size=12, color=ACCENT), fmt=fmt,
        fill=EDIT, align=Alignment(horizontal="right"))
    put(cfg, f"C{i}", note, font=Font(size=9, color=MUTED), align=Alignment(indent=1))
    cfg.row_dimensions[i].height = 22
put(cfg, "A8", "美元汇率(自动)", font=Font(size=11, color=INK), align=Alignment(indent=1))
put(cfg, "B8", '=IFERROR(GOOGLEFINANCE("CURRENCY:USDJPY"),160.06)',
    font=Font(bold=True, size=12, color=ACCENT), fmt='#,##0.00',
    align=Alignment(horizontal="right"))
put(cfg, "C8", "GOOGLEFINANCE 自动取。取不到就退回 160.06（2026/08/29 的值）。",
    font=Font(size=9, color=MUTED), align=Alignment(indent=1))
cfg.row_dimensions[8].height = 22
put(cfg, "A10", "黄色格子随便改。其他所有表都引用这里。",
    font=Font(size=9, italic=True, color=MUTED))

# -------------------------------------------------------------- 持仓明细 --
hold = wb.create_sheet(S_HOLD)
hold.sheet_view.showGridLines = False
put(hold, "A1", "持仓明细 —— 每月更新「评价额」一列即可",
    font=Font(bold=True, size=14, color=INK))
hold.merge_cells("A1:I1"); hold.row_dimensions[1].height = 26
header(hold, 2, ["名称", "类别", "账户", "币种", "代码", "数量", "价格倍率",
                 "实时价(自动)", "价格(手动/CSV)", "评价额(日元)", "成本(日元)",
                 "盈亏", "占比", "备注", "CSV銘柄", "CSV口座"],
       (28, 11, 19, 7, 15, 13, 9, 13, 15, 15, 14, 14, 8, 26, 34, 19))
hold.freeze_panes = "B3"

dv_cat = DataValidation(type="list", formula1='"' + ",".join(CATS) + '"', allow_blank=True)
dv_acc = DataValidation(type="list", formula1='"' + ",".join(ACCTS) + '"', allow_blank=True)
hold.add_data_validation(dv_cat); dv_cat.add(f"B3:B{HOLDS+2}")
hold.add_data_validation(dv_acc); dv_acc.add(f"C3:C{HOLDS+2}")


def csv_aware(r, expr, fallback):
    """CSV に一致行があればそれを使い、無ければ 2026/08/29 の値を残す。"""
    hit = f'COUNTIFS({C_NAME},$O{r},{C_ACC},$P{r})'
    return f'=IF($O{r}="",{fallback},IF({hit}=0,{fallback},{expr}))'


CASH_HIT = f'COUNTIF({C_LBL},"預り金")'
CASH_SUM = (f'SUMIF({C_LBL},"預り金",{C_LVAL})+SUMIF({C_LBL},"信用保証金",{C_LVAL})'
            f'+SUMIF({C_LBL},"外貨預り金",{C_LVAL})')

for i, (nm, cat, acc, cur, tick, qty, mul, px, cost, cnm, cac) in enumerate(HOLDINGS):
    r = 3 + i
    for col, v in zip("ABCDEG", (nm, cat, acc, cur, tick, mul)):
        hold[f"{col}{r}"] = v
    hold[f"O{r}"], hold[f"P{r}"] = cnm, cac

    if cnm == "@CASH":          # 預り金はヘッダーブロック側にあるので別扱い
        hold[f"F{r}"] = qty
        hold[f"I{r}"] = f'=IF({CASH_HIT}=0,{px},{CASH_SUM})'
        hold[f"K{r}"] = f'=IF({CASH_HIT}=0,{cost},{CASH_SUM})'
    elif cnm:
        hold[f"F{r}"] = csv_aware(r, f'SUMIFS({C_QTY},{C_NAME},$O{r},{C_ACC},$P{r})', qty)
        hold[f"I{r}"] = csv_aware(r, f'AVERAGEIFS({C_PX},{C_NAME},$O{r},{C_ACC},$P{r})', px)
        hold[f"K{r}"] = csv_aware(
            r, f'SUMIFS({C_VAL},{C_NAME},$O{r},{C_ACC},$P{r})'
               f'-SUMIFS({C_PL},{C_NAME},$O{r},{C_ACC},$P{r})', cost)
    else:                        # 楽天以外は手入力のまま
        hold[f"F{r}"] = qty
        if px is not None:   hold[f"I{r}"] = px
        if cost is not None: hold[f"K{r}"] = cost

for r in range(3, HOLDS + 3):
    hold[f"H{r}"] = f'=IF($E{r}="","",IFERROR(GOOGLEFINANCE($E{r}),""))'
    hold[f"J{r}"] = (f'=IF($F{r}="","",$F{r}/$G{r}*IF(ISNUMBER($H{r}),$H{r},$I{r})'
                     f'*IF($D{r}="{USD}",{FX},1))')
    hold[f"L{r}"] = f'=IF(OR($J{r}="",$K{r}=""),"",$J{r}-$K{r})'
    hold[f"M{r}"] = f'=IF($J{r}="","",IFERROR($J{r}/SUM({H_VAL}),""))'
    hold[f"N{r}"] = f'=IF($A{r}="","",IF($J{r}=0,"← 填价格，总资产才准",""))'
    for col in "ABCDEG":
        hold[f"{col}{r}"].fill = PatternFill("solid", fgColor=EDIT)
    for col, f in (("F", '#,##0.######'), ("H", '#,##0.00'), ("I", '#,##0.00'),
                   ("J", YEN), ("K", YEN), ("L", YEN), ("M", PCT)):
        hold[f"{col}{r}"].number_format = f
    for col in "FHIKLM":
        hold[f"{col}{r}"].font = Font(size=10, color=MUTED)
    hold[f"J{r}"].font = Font(bold=True, size=10, color=INK)
    hold[f"N{r}"].font = Font(size=9, bold=True, color=BAD)
    for col in "OP":
        hold[f"{col}{r}"].font = Font(size=8, color="94A3B8")
hold.conditional_formatting.add(
    f"L3:L{HOLDS+2}", CellIsRule(operator="lessThan", formula=["0"], font=Font(color=BAD)))

for i, t in enumerate([
    "评价额 = 数量 ÷ 价格倍率 × 价格 × 汇率。「实时价」有数就用它，没有才用「价格(手动/CSV)」。",
    "数量 / 价格 / 成本 三列会自动认 CSV 页 —— 贴了新的 assetbalanceall 就自动更新，没贴就保持 2026/08/29 的快照。",
    "最右边两列是跟 CSV 对照用的键（楽天写法的銘柄名和口座）。楽天改了名字才需要动它。",
    "美股和汇率走 GOOGLEFINANCE 实时。日本股的 TYO: 代码目前取不到值，会自动退回 CSV 的价格。",
    "BTC / 野村 / ゆうちょ 不在 CSV 里，永远手填。",
]):
    put(hold, f"A{HOLDS+4+i}", "・" + t, font=Font(size=9, color=MUTED))

# ------------------------------------------------------------------- CSV --
csvs = wb.create_sheet(S_CSV)
csvs.column_dimensions["A"].width = 22
csvs.column_dimensions["C"].width = 46
csvs.column_dimensions["D"].width = 20
put(csvs, "A1", "这一页整页都是给 assetbalanceall CSV 用的。下面的字会被覆盖掉，没关系。",
    font=Font(bold=True, size=12, color=INK))
for i, t in enumerate([
    "",
    "更新步骤（每月一次，约 1 分钟）：",
    "  1. 楽天証券 → 保有商品一覧 → CSVダウンロード（文件名形如 assetbalanceall_20260829_231124.csv）",
    "  2. 先点一下这一页的标签，确保当前在「CSV」页",
    "  3. 文件 > 导入 > 上传 > 选那个 CSV",
    "  4. 导入位置选「替换当前工作表」← 一定要选这个，别选「替换电子表格」",
    "  5. 分隔符选「逗号」，然后导入",
    "",
    "完成。持仓明细的数量 / 价格 / 成本会自动跟着变，你不用动任何格子。",
    "",
    "没导入之前，表里用的是 2026/08/29 那份快照的数字。",
]):
    put(csvs, f"A{2+i}", t, font=Font(size=10, color=MUTED))
# -------------------------------------------------------------- 月次记录 --
mon = wb.create_sheet(S_MON)
mon.sheet_view.showGridLines = False
put(mon, "A1", "月次记录 —— 每月两个数：总资产、这个月投了多少",
    font=Font(bold=True, size=14, color=INK))
mon.merge_cells("A1:G1"); mon.row_dimensions[1].height = 26
header(mon, 2, ["年月", "总资产", "本月投入", "累计本金", "评价盈亏", "达成", "连续"],
       (13, 16, 15, 16, 16, 8, 8))
mon.freeze_panes = "A3"

for i, (ym, amt) in enumerate(PAST):
    r = 3 + i
    y, m = ym.split("-")
    mon[f"A{r}"] = f"=DATE({y},{m},1)"
    mon[f"C{r}"] = amt

for r in range(3, MONTHS + 3):
    a = f"$A{r}"
    mon[f"D{r}"] = f'=IF({a}="","",{BASE}+SUMIFS({M_IN},{M_YM},"<="&{a}))'
    mon[f"E{r}"] = f'=IF(OR({a}="",$B{r}=""),"",$B{r}-$D{r})'
    mon[f"F{r}"] = f'=IF({a}="","",IF($C{r}>={TARGET},"✓","–"))'
    mon[f"G{r}"] = (f'=IF({a}="","",IF($C{r}>={TARGET},1,0))' if r == 3 else
                    f'=IF({a}="","",IF($C{r}>={TARGET},IF($G{r-1}="",1,$G{r-1}+1),0))')
    mon[f"A{r}"].number_format = MON
    for col in "BCDE":
        mon[f"{col}{r}"].number_format = YEN
    for col in "ABC":
        mon[f"{col}{r}"].fill = PatternFill("solid", fgColor=EDIT)
    for col in "DE":
        mon[f"{col}{r}"].font = Font(size=10, color=MUTED)
    mon[f"B{r}"].font = Font(bold=True, size=10, color=INK)
    for col in "FG":
        mon[f"{col}{r}"].alignment = Alignment(horizontal="center")
mon.conditional_formatting.add(
    f"C3:C{MONTHS+2}", CellIsRule(operator="lessThan", formula=["0"], font=Font(bold=True, color=BAD)))
mon.conditional_formatting.add(
    f"E3:E{MONTHS+2}", CellIsRule(operator="lessThan", formula=["0"], font=Font(color=BAD)))
mon.column_dimensions["I"].width = 52
put(mon, "I3", "「本月投入」这12行是从你的取引履歴算出来的真实净入金。",
    font=Font(size=10, bold=True, color=ACCENT))
put(mon, "I4", "负数＝那个月你从楽天証券取走的钱比投进去的多。",
    font=Font(size=10, color=BAD))
put(mon, "I5", "「总资产」留空是故意的：先去持仓明细填上 BTC / 野村 / ゆうちょ，",
    font=Font(size=10, color=MUTED))
put(mon, "I6", "然后把总览里那个自动算好的总资产抄一个数过来。",
    font=Font(size=10, color=MUTED))

# ------------------------------------------------------------------ 总览 --
d = wb.create_sheet(S_DASH, 0)
d.sheet_view.showGridLines = False
for c, w in zip("ABCDEFG", (2, 24, 20, 16, 24, 3, 3)):
    d.column_dimensions[c].width = w

TOTAL = "$C$6"; THIS = "$C$14"; PACE = "$C$20"; MP = "$C$21"; MT = "$C$22"
NEXT = "$C$45"
LAST = lambda col: (
    f"IFERROR(INDEX('{S_MON}'!${col}$3:${col}${MONTHS+2},"
    f"MATCH(MAX({M_YM}),{M_YM},0)),0)")

huge = Font(bold=True, size=26, color=ACCENT)
val  = Font(bold=True, size=12, color=INK)
sub  = Font(size=10, color=MUTED)
barf = Font(size=11, color=ACCENT)

put(d, "B2", "1亿日元计划", font=Font(bold=True, size=24, color=INK))
d.row_dimensions[2].height = 36
put(d, "B3", f'="目标 "&TEXT({GOAL},"#,##0")&"円 ／ 每月 "&TEXT({TARGET},"#,##0")'
             f'&"円 ／ 预期年化 "&TEXT({RATE},"0.0%")', font=sub)
put(d, "E2", "=TODAY()", font=sub, fmt=DAY, align=Alignment(horizontal="right"))

band(d, 5, "  现在")
label(d, 6, "总资产");  put(d, "C6", f"=SUM({H_VAL})", font=huge, fmt=YEN)
d.row_dimensions[6].height = 34
put(d, "D6", '="（持仓明细自动合计）"', font=sub)
put(d, "E6", f'=IF(COUNTIF({C_LBL},"種別")=0,"CSV 未导入 —— 用的是 2026/08/29 快照",'
             f'"CSV 已导入 "&IFERROR(INDEX({C_ACC},MATCH("米ドル",{C_LBL},0)),""))',
    font=Font(size=9, italic=True, color=MUTED))
label(d, 7, "距离目标"); put(d, "C7", f"={GOAL}-{TOTAL}", font=val, fmt=YEN)
label(d, 8, "进度");     put(d, "C8", f"={TOTAL}/{GOAL}", font=val, fmt=PCT)
put(d, "E8", bar(f"{TOTAL}/{GOAL}"), font=barf)
label(d, 9, "累计本金"); put(d, "C9", f"={LAST('D')}", font=val, fmt=YEN)
label(d, 10, "浮动盈亏")
put(d, "C10", f"=IF($C$9=0,0,{TOTAL}-$C$9)", font=val, fmt='+#,##0"円";-#,##0"円"')
put(d, "D10", "=IFERROR($C$10/$C$9,0)", font=sub, fmt='+0.0%;-0.0%')
put(d, "E10", '="设置里的起始本金要含野村/ゆうちょ/BTC，这个数才准"',
    font=Font(size=9, italic=True, color=BAD))
for op, col in (("lessThan", BAD), ("greaterThan", GOOD)):
    d.conditional_formatting.add("C10:D10",
        CellIsRule(operator=op, formula=["0"], font=Font(bold=True, color=col)))

band(d, 13, '="  本月 ── "&TEXT(TODAY(),"yyyy年m月")')
label(d, 14, "已投入")
put(d, "C14", f'=SUMIFS({M_IN},{M_YM},">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1),'
              f'{M_YM},"<"&EOMONTH(TODAY(),0)+1)', font=huge, fmt=YEN)
d.row_dimensions[14].height = 34
put(d, "D14", f'=" / "&TEXT({TARGET},"#,##0")&"円"', font=sub)
label(d, 15, "达成率"); put(d, "C15", f"=IFERROR({THIS}/{TARGET},0)", font=val, fmt=PCT)
put(d, "E15", bar(f"{THIS}/{TARGET}"), font=barf)
label(d, 16, "还差");   put(d, "C16", f"=MAX(0,{TARGET}-{THIS})", font=val, fmt=YEN)
put(d, "D16", '="这个月还剩"&(EOMONTH(TODAY(),0)-TODAY())&"天"', font=sub)
label(d, 17, "连续达标"); put(d, "C17", f'={LAST("G")}&"个月"', font=val)
d.conditional_formatting.add("C15",
    CellIsRule(operator="greaterThanOrEqual", formula=["1"], font=Font(bold=True, color=GOOD)))
d.conditional_formatting.add("C15",
    CellIsRule(operator="lessThan", formula=["0.5"], font=Font(bold=True, color=BAD)))

band(d, 19, "  什么时候能到")
label(d, 20, "现在的速度")
put(d, "C20", f'=IFERROR(IF(COUNTIFS({M_YM},">="&EDATE(TODAY(),-12))>=3,'
              f'AVERAGEIFS({M_IN},{M_YM},">="&EDATE(TODAY(),-12)),{PACE0}),{PACE0})',
    font=val, fmt=YEN)
put(d, "E20", '="（近12个月平均，自动滚动）"', font=Font(size=9, italic=True, color=MUTED))
put(d, "D20", '="／月"', font=sub)
label(d, 21, "  按这个速度")
put(d, "C21", f'=IFERROR(NPER({MRATE},-{PACE},-{TOTAL},{GOAL}),"")', font=val, fmt='0.0')
put(d, "D21", mlabel(MP), font=Font(bold=True, size=13, color=INK))
put(d, "E21", f'=IFERROR(TEXT(EDATE(TODAY(),ROUNDUP({MP},0)),"yyyy年m月")&" 到达","—")', font=sub)
label(d, 22, "  按计划40万")
put(d, "C22", f'=IFERROR(NPER({MRATE},-{TARGET},-{TOTAL},{GOAL}),"")', font=val, fmt='0.0')
put(d, "D22", mlabel(MT), font=Font(bold=True, size=13, color=GOOD))
put(d, "E22", f'=IFERROR(TEXT(EDATE(TODAY(),ROUNDUP({MT},0)),"yyyy年m月")&" 到达","—")', font=sub)
label(d, 23, "差距")
put(d, "D23", mlabel(f"MAX(0,{MP}-{MT})"), font=Font(bold=True, size=18, color=BAD))
d.row_dimensions[23].height = 28
put(d, "E23", '="每月少投的那部分，变成了你的时间"', font=Font(size=9, italic=True, color=MUTED))
label(d, 24, "整月不投的话")
put(d, "D24", f'=IFERROR("目标推迟约"&TEXT(NPER({MRATE},-{TARGET},'
              f'-({TOTAL}*(1+{MRATE})),{GOAL})+1-{MT},"0.0")&"个月","—")',
    font=Font(bold=True, size=11, color=BAD))
label(d, 25, "每月多投10万")
put(d, "D25", f'=IFERROR("目标提前约"&TEXT({MT}-NPER({MRATE},-({TARGET}+100000),'
              f'-{TOTAL},{GOAL}),"0.0")&"个月","—")', font=Font(bold=True, size=11, color=GOOD))

band(d, 28, "  资产构成（条形按 33% 满格）")
for i, cat in enumerate(CATS):
    r = 29 + i
    put(d, f"B{r}", cat, font=Font(size=10, color=INK), align=Alignment(indent=1))
    put(d, f"C{r}", f'=SUMIF({H_CAT},$B{r},{H_VAL})', font=Font(size=10, color=INK), fmt=YEN)
    put(d, f"D{r}", f'=IFERROR($C{r}/{TOTAL},0)', font=sub, fmt=PCT)
    put(d, f"E{r}", bar(f"$C{r}/{TOTAL}*3", 18), font=Font(size=10, color=ACCENT))

band(d, 40, "  NISA 生涯投资额度")
label(d, 41, "已经用掉")
put(d, "C41", f'=SUMIF({H_ACC},"NISA*",{H_CST})', font=val, fmt=YEN)
put(d, "D41", "=IFERROR($C$41/18000000,0)", font=sub, fmt=PCT)
put(d, "E41", bar("$C$41/18000000"), font=barf)
label(d, 42, "还剩")
put(d, "C42", "=MAX(0,18000000-$C$41)", font=val, fmt=YEN)
put(d, "D42", '="额度按取得原价计算，不是按评价额"', font=Font(size=9, italic=True, color=MUTED))

band(d, 44, "  下一个里程碑")
label(d, 45, "下一档")
put(d, "C45", f"=IF({TOTAL}<15000000,15000000,IF({TOTAL}<20000000,20000000,"
              f"IF({TOTAL}<30000000,30000000,IF({TOTAL}<50000000,50000000,"
              f"IF({TOTAL}<70000000,70000000,{GOAL})))))",
    font=Font(bold=True, size=16, color=ACCENT), fmt=YEN)
label(d, 46, "还差")
put(d, "C46", f"={NEXT}-{TOTAL}", font=val, fmt=YEN)
put(d, "E46", bar(f"{TOTAL}/{NEXT}"), font=barf)
label(d, 47, "按现在速度")
put(d, "C47", f'=IFERROR(NPER({MRATE},-{PACE},-{TOTAL},{NEXT}),"")', font=val, fmt='0.0')
put(d, "D47", mlabel("$C$47"), font=Font(bold=True, size=13, color=INK))

band(d, 50, "  别忘了")
for i, t in enumerate([
    "84% 的资产在 NISA 里，356万含み益目前免税。NISA 卖出后额度要等次年才恢复 —— 别乱动。",
    "美国资产约占 78%（含オルカン的美国部分）。日元升值会直接吃掉收益，这是最大的单一风险。",
    "黄金 14%。它不产生现金流，是保险，不是引擎。别指望它把你送到1亿。",
    "個別股 17%：AMD +290%、GOOG +107%，但メイコー -48%。投信（69%）才是主力。",
    "评价额会上下跳，你能控制的只有「累计本金」。看那一条。",
    "BTC 数量 0.061912、成本 893,703 来自你自己的「Bitcoin trade history」表，最后一笔 2026/01/29。之后有交易的话要自己改。",
]):
    put(d, f"B{51+i}", "・" + t, font=Font(size=9, color=MUTED), align=Alignment(indent=1))
    d.merge_cells(f"B{51+i}:E{51+i}")

c1 = LineChart(); c1.title = "本金 vs 总资产"; c1.height, c1.width = 7.5, 15
c1.y_axis.numFmt = '#,##0'
c1.add_data(Reference(mon, min_col=2, max_col=2, min_row=2, max_row=MONTHS + 2), titles_from_data=True)
c1.add_data(Reference(mon, min_col=4, max_col=4, min_row=2, max_row=MONTHS + 2), titles_from_data=True)
c1.set_categories(Reference(mon, min_col=1, min_row=3, max_row=MONTHS + 2))
d.add_chart(c1, "H3")

c2 = BarChart(); c2.title = "每月投入"; c2.height, c2.width = 7.5, 15
c2.y_axis.numFmt = '#,##0'
c2.add_data(Reference(mon, min_col=3, max_col=3, min_row=2, max_row=MONTHS + 2), titles_from_data=True)
c2.set_categories(Reference(mon, min_col=1, min_row=3, max_row=MONTHS + 2))
d.add_chart(c2, "H22")

wb.save(sys.argv[1])
print("wrote", sys.argv[1])
