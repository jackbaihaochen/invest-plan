# -*- coding: utf-8 -*-
"""生成した .xlsx の数式を機械的に検査する。

LibreOffice がこの環境で xlsx を開けないため、評価の代わりに
構文・関数名・自己参照・シート名参照を検査している。
"""
import re, sys
from openpyxl import load_workbook

ALLOWED = {  # Excel と Google スプレッドシートの両方に存在する関数だけ
 "IF","IFERROR","SUM","SUMIFS","AVERAGEIF","COUNTIF","COUNT","MIN","MAX","ROUND",
 "ROUNDUP","INT","MOD","TEXT","DATE","YEAR","MONTH","TODAY","EOMONTH","EDATE",
 "REPT","LOOKUP","NPER","ROW",
}

wb = load_workbook(sys.argv[1])
names = set(wb.sheetnames)
bad = []
funcs = set()

def strip_strings(f):
    out, ins = [], False
    for ch in f:
        if ch == '"':
            ins = not ins
            continue
        out.append(" " if ins else ch)
    return "".join(out), ins

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            f = c.value
            if not isinstance(f, str) or not f.startswith("="):
                continue
            where = f"{ws.title}!{c.coordinate}"
            body, unterminated = strip_strings(f)
            if unterminated:
                bad.append((where, "引用符が閉じていない", f))
            if body.count("(") != body.count(")"):
                bad.append((where, "括弧が合っていない", f))
            for fn in re.findall(r"([A-Z][A-Z0-9\.]*)\s*\(", body):
                funcs.add(fn)
                if fn not in ALLOWED:
                    bad.append((where, f"未確認の関数 {fn}", f))
            for sh in re.findall(r"'([^']+)'!", f):
                if sh not in names:
                    bad.append((where, f"存在しないシート {sh}", f))
            # 自己参照（絶対参照で自分自身を指していないか）
            if re.search(rf"\$?{c.column_letter}\$?{c.row}\b", body) and "!" not in body:
                bad.append((where, "自己参照の疑い", f))

print("使用関数:", " ".join(sorted(funcs)))
print("シート:", " / ".join(wb.sheetnames))
if bad:
    print(f"\n❌ {len(bad)} 件")
    seen = set()
    for w, why, f in bad:
        k = (why, f[:60])
        if k in seen:
            continue
        seen.add(k)
        print(f"  {w}: {why}\n     {f[:150]}")
    sys.exit(1)
print("\n✅ 数式チェック通過")
